#!/usr/bin/env python3
"""
Scans Jira projects and issues for exposed secrets.
Authenticates via Atlassian email and API token.
Exports findings to a formatted Excel report.
"""

import requests
from requests.auth import HTTPBasicAuth
import json
import sys
import argparse
import os
import re
import shutil
import time
import signal
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings

import yaml

# Progress bar support (optional)
try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

# OCR support (optional)
try:
    from PIL import Image
    import pytesseract
    os.environ["TESSDATA_PREFIX"] = "/usr/local/share/tessdata/"
    warnings.simplefilter('ignore', Image.DecompressionBombWarning)
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# PDF support (optional)
try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# DOCX support (optional)
try:
    import docx as python_docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# AWS SES for email delivery
try:
    import boto3
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    AWS_SES_AVAILABLE = True
except ImportError:
    AWS_SES_AVAILABLE = False

# ─────────────────────────────────────────────
# Global state for graceful Ctrl+C handling (#3)
# ─────────────────────────────────────────────
_partial_findings: List[Dict] = []
_interrupted = False


def _handle_sigint(signum, frame):
    """Handle Ctrl+C: set flag so main loop can save partial results."""
    global _interrupted
    _interrupted = True
    print("\n\n⚠️  Interrupted! Saving partial results…")


signal.signal(signal.SIGINT, _handle_sigint)


# ─────────────────────────────────────────────
# Retry / rate-limit decorator (#5)
# ─────────────────────────────────────────────
def with_retry(max_retries: int = 3, backoff: float = 2.0):
    """
    Decorator that retries a function returning a requests.Response on
    HTTP 429 (rate-limit) or transient network errors, with exponential
    back-off.
    """
    def decorator(func):
        from functools import wraps

        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    result = func(*args, **kwargs)
                    # If it's a Response and we're being rate-limited, back off
                    if hasattr(result, 'status_code') and result.status_code == 429:
                        retry_after = float(result.headers.get('Retry-After', backoff ** (attempt + 1)))
                        print(f"⏳ Rate-limited (429). Waiting {retry_after:.1f}s before retry {attempt + 1}/{max_retries}…")
                        time.sleep(retry_after)
                        continue
                    return result
                except requests.exceptions.RequestException as exc:
                    if attempt == max_retries - 1:
                        raise
                    wait = backoff ** (attempt + 1)
                    print(f"⚠️  Network error: {exc}. Retrying in {wait:.1f}s ({attempt + 1}/{max_retries})…")
                    time.sleep(wait)
            return None  # exhausted retries
        return wrapper
    return decorator


# ─────────────────────────────────────────────
# .env loader
# ─────────────────────────────────────────────
def load_env_file(env_path='.env'):
    """Load variables from a .env file."""
    env_vars = {}
    env_file = Path(env_path)

    if env_file.exists():
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_vars[key.strip()] = value.strip()

    return env_vars


def normalize_jira_url(url):
    """Normalize Jira URL by stripping trailing slash."""
    return url.rstrip('/')


# ─────────────────────────────────────────────
# Pattern loading with compiled regex cache (#6)
# ─────────────────────────────────────────────
def load_secret_patterns(patterns_file='secret_patterns.txt') -> List[Tuple[str, re.Pattern, int]]:
    """
    Load secret detection patterns from file.
    Format: Name:::Regex:::GroupIndex

    Returns compiled patterns: List[(name, compiled_regex, group_index)]
    """
    raw_patterns = []

    if not Path(patterns_file).exists():
        print(f"⚠️  Patterns file {patterns_file} not found. Using built-in fallback patterns.")
        raw_patterns = [
            ('AWS Access Key ID', r'(?:^|[^A-Za-z0-9])((AKIA|ASIA|AGPA|AIDA|AROA|AIPA|ANPA|ANVA)(?!([A-Z0-9])\3{5,})[A-Z0-9]{16})(?:[^A-Za-z0-9]|$)', 1),
            ('GitHub Personal Access Token', r'(?:^|[^a-z0-9_])(ghp_[0-9a-zA-Z]{36})(?:[^a-zA-Z0-9]|$)', 1),
            ('Atlassian API Token', r'(?:^|[^A-Z])(ATATT[a-zA-Z0-9\-_]{28,})(?:[^a-zA-Z0-9\-_]|$)', 1),
        ]
    else:
        with open(patterns_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and ':::' in line:
                    parts = line.split(':::')
                    if len(parts) >= 3:
                        name = parts[0].strip()
                        regex = parts[1].strip()
                        group_index = int(parts[2].strip())
                        raw_patterns.append((name, regex, group_index))

    # Compile once (#6)
    compiled = []
    for name, regex, group_index in raw_patterns:
        try:
            compiled.append((name, re.compile(regex, re.MULTILINE | re.IGNORECASE), group_index))
        except re.error as e:
            print(f"⚠️  Invalid regex pattern '{name}': {e}")
    return compiled


def load_trufflehog_patterns(patterns_file, include_keywords=None, exclude_keywords=None):
    """
    Loads TruffleHog v3 YAML patterns and converts them to internal format.
    Returns (compiled_patterns, stats) always — even on error.
    """
    _empty_stats = {'total': 0, 'skipped_include': 0, 'skipped_exclude': 0, 'total_skipped': 0, 'loaded': 0}

    if not Path(patterns_file).exists():
        print(f"⚠️  TruffleHog patterns file '{patterns_file}' not found.")
        return [], _empty_stats

    try:
        with open(patterns_file, 'r', encoding='utf-8') as f:
            rules = yaml.safe_load(f)
    except yaml.YAMLError as e:
        print(f"⚠️  Failed to parse TruffleHog YAML file: {e}")
        return [], _empty_stats

    if not isinstance(rules, list):
        print(f"⚠️  Expected a list of rules in '{patterns_file}', got {type(rules).__name__}")
        return [], _empty_stats

    include_set = {kw.strip().lower() for kw in include_keywords} if include_keywords else None
    exclude_set = {kw.strip().lower() for kw in exclude_keywords} if exclude_keywords else None

    skipped_include = 0
    skipped_exclude = 0
    patterns = []

    for rule in rules:
        name = rule.get('name', 'unknown')
        regexes = rule.get('regex', {})
        rule_keywords = [kw.lower() for kw in rule.get('keywords', [])]

        if not isinstance(regexes, dict):
            continue

        if include_set is not None:
            if not any(kw in include_set for kw in rule_keywords):
                skipped_include += 1
                continue

        if exclude_set is not None:
            if any(kw in exclude_set for kw in rule_keywords):
                skipped_exclude += 1
                continue

        for regex_name, regex_pattern in regexes.items():
            if not regex_pattern:
                continue
            try:
                compiled = re.compile(regex_pattern, re.MULTILINE | re.IGNORECASE)  # compiled once (#6)
                group_index = 1 if compiled.groups > 0 else 0
                patterns.append((name, compiled, group_index))
            except re.error as e:
                print(f"⚠️  Invalid regex in TruffleHog rule '{name}': {e}")
                continue

    total_skipped = skipped_include + skipped_exclude
    stats = {
        'total':           len(rules),
        'skipped_include': skipped_include,
        'skipped_exclude': skipped_exclude,
        'total_skipped':   total_skipped,
        'loaded':          len(patterns),
    }
    return patterns, stats


# ─────────────────────────────────────────────
# Whitelist / false-positive filter (#18)
# ─────────────────────────────────────────────
def load_ignore_list(ignore_file: str) -> set:
    """
    Load a whitelist of known false-positives from file.

    Format (one entry per line, comments with #):
        ISSUE-123:AWS Access Key ID:AKIAIOSFODNN7EXAMPLE
        *:GitHub Personal Access Token:ghp_testtoken12345678901234567890

    Returns a set of (issue_key_or_*, secret_type, secret_value) tuples.
    Wildcards: use '*' for issue_key to ignore a value everywhere.
    """
    ignore_set = set()
    path = Path(ignore_file)

    if not path.exists():
        print(f"⚠️  Ignore file '{ignore_file}' not found — no whitelist applied.")
        return ignore_set

    with open(path, 'r', encoding='utf-8') as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = line.split(':', 2)
            if len(parts) != 3:
                print(f"⚠️  Ignore file line {lineno} malformed (expected issue:type:value): {line!r}")
                continue
            issue_key, secret_type, secret_value = (p.strip() for p in parts)
            ignore_set.add((issue_key, secret_type, secret_value))

    print(f"✅ Ignore list loaded: {len(ignore_set)} entr{'y' if len(ignore_set) == 1 else 'ies'}")
    return ignore_set


def is_ignored(finding: Dict, ignore_set: set) -> bool:
    """Return True if finding matches any entry in the ignore set."""
    if not ignore_set:
        return False
    issue_key = finding.get('issue_key', '')
    secret_type = finding.get('secret_type', '')
    secret_value = finding.get('secret_value', '')
    # Exact match
    if (issue_key, secret_type, secret_value) in ignore_set:
        return True
    # Wildcard match (any issue)
    if ('*', secret_type, secret_value) in ignore_set:
        return True
    return False


# ─────────────────────────────────────────────
# Incremental scan state (#17)
# ─────────────────────────────────────────────
_STATE_FILE = '.jira_scanner_state.json'


def load_scan_state(state_file: str = _STATE_FILE) -> Dict:
    """Load previous scan state (last scan timestamps per project)."""
    path = Path(state_file)
    if path.exists():
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def save_scan_state(state: Dict, state_file: str = _STATE_FILE):
    """Persist scan state to disk."""
    with open(state_file, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=2)


# ─────────────────────────────────────────────
# Secret scanning
# ─────────────────────────────────────────────
def scan_text_for_secrets(text: str, patterns: List[Tuple[str, re.Pattern, int]]) -> List[Dict]:
    """
    Scan text for secrets using pre-compiled regex patterns.
    Accepts (name, compiled_pattern, group_index) tuples.
    """
    findings = []

    for pattern_name, compiled_pattern, group_index in patterns:
        try:
            for match in compiled_pattern.finditer(text):
                if group_index < len(match.groups()) + 1:
                    secret_value = match.group(group_index) if group_index > 0 else match.group(0)
                    start = max(0, match.start() - 50)
                    end = min(len(text), match.end() + 50)
                    context = text[start:end].replace('\n', ' ').replace('\r', '').strip()
                    findings.append({
                        'secret_type': pattern_name,
                        'secret_value': secret_value,
                        'context': context
                    })
        except re.error as e:
            print(f"⚠️  Regex error in pattern '{pattern_name}': {e}")
            continue

    return findings


def get_issue_attachments(email, api_token, jira_url, issue_key):
    """Fetch the list of attachments for a given Jira issue."""
    jira_url = normalize_jira_url(jira_url)
    url = f"{jira_url}/rest/api/2/issue/{issue_key}?fields=attachment"
    auth = HTTPBasicAuth(email, api_token)
    headers = {"Accept": "application/json"}

    try:
        response = requests.get(url, headers=headers, auth=auth, timeout=30)
        if response.status_code == 200:
            fields = response.json().get("fields", {})
            return fields.get("attachment", [])
    except Exception as e:
        print(f"⚠️  Failed to fetch attachments for {issue_key}: {e}")
    return []


def extract_text_from_attachment(attachment, email, api_token, max_size_bytes=None):
    """Download an attachment and extract text from it."""
    att_title = attachment.get("filename", "unknown")
    download_url = attachment.get("content", "")
    file_size = attachment.get("size", 0)
    ext = os.path.splitext(att_title)[1].lstrip(".").lower()

    if max_size_bytes and file_size > max_size_bytes:
        return "", ext

    supported_text_exts = {
        "txt", "log", "md", "rst", "conf", "cfg", "ini", "properties", "env",
        "json", "xml", "yaml", "yml", "toml", "csv", "tsv",
        "py", "sh", "bash", "bat", "ps1", "js", "ts", "java", "go", "rb",
        "php", "cs", "c", "cpp", "h", "sql", "tf", "hcl", "dockerfile",
        "html", "htm", "css",
    }
    supported_image_exts = {"png", "jpg", "jpeg", "gif", "bmp", "tiff"}

    is_text = ext in supported_text_exts
    is_image = ext in supported_image_exts
    is_pdf = ext == "pdf"
    is_docx = ext == "docx"

    if not (is_text or is_image or is_pdf or is_docx):
        return "", ext

    try:
        auth = HTTPBasicAuth(email, api_token)
        response = requests.get(download_url, auth=auth, timeout=30)
        response.raise_for_status()
        content = response.content

        if is_text:
            return content.decode("utf-8", errors="ignore"), ext
        if is_docx and DOCX_AVAILABLE:
            doc = python_docx.Document(BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs), ext
        if is_pdf and PDF_AVAILABLE:
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            return "\n".join(page.get_text() for page in pdf_doc), ext
        if is_image and OCR_AVAILABLE:
            img = Image.open(BytesIO(content))
            return pytesseract.image_to_string(img), ext
    except Exception as e:
        print(f"⚠️  Error extracting text from attachment '{att_title}': {e}")

    return "", ext


# ─────────────────────────────────────────────
# Jira API helpers (with retry decorator) (#5)
# ─────────────────────────────────────────────
@with_retry(max_retries=3, backoff=2.0)
def _get(url, headers, auth, params=None, timeout=30):
    return requests.get(url, headers=headers, auth=auth, params=params, timeout=timeout)


def get_jira_projects(email, api_token, jira_url):
    """Fetch all accessible Jira projects with full details."""
    jira_url = normalize_jira_url(jira_url)
    url = f"{jira_url}/rest/api/3/project/search"
    auth = HTTPBasicAuth(email, api_token)
    params = {'expand': 'description,lead,url,insight'}
    headers = {"Accept": "application/json"}

    all_projects = []
    start_at = 0
    max_results = 50

    try:
        while True:
            params['startAt'] = start_at
            params['maxResults'] = max_results

            response = _get(url, headers, auth, params=params)

            if response is None:
                print("❌ Request failed after retries.")
                return None
            if response.status_code == 200:
                data = response.json()
                projects = data.get('values', [])
                all_projects.extend(projects)
                if data.get('isLast', True):
                    break
                start_at += max_results
            elif response.status_code == 401:
                print("❌ Authentication failed. Check your email and API token.")
                return None
            elif response.status_code == 403:
                print("❌ Access denied. Check your permissions.")
                return None
            else:
                print(f"❌ Error: {response.status_code}")
                print(f"Response: {response.text}")
                return None

        return all_projects
    except requests.exceptions.RequestException as e:
        print(f"❌ Request failed: {e}")
        return None


def get_project_issues(email, api_token, jira_url, project_key,
                       max_issues=0, verbose=False, since_date: Optional[str] = None,
                       created_after: Optional[str] = None):
    """
    Fetch issues for a given project.
    Tries multiple API endpoints for compatibility.
    max_issues=0 means fetch all issues.

    Args:
        since_date:    ISO date string — only fetch issues updated on or after this date.
        created_after: ISO date string — only fetch issues created on or after this date.
    """
    jira_url = normalize_jira_url(jira_url)
    auth = HTTPBasicAuth(email, api_token)
    headers = {"Accept": "application/json"}

    all_issues = []
    unlimited = (max_issues == 0)
    effective_limit = 999999 if unlimited else max_issues

    if verbose:
        limit_text = "all" if unlimited else str(max_issues)
        filters = []
        if since_date:
            filters.append(f"updated >= {since_date}")
        if created_after:
            filters.append(f"created >= {created_after}")
        filter_text = (", " + ", ".join(filters)) if filters else ""
        print(f"   🔍 Fetching issues for project {project_key} (limit: {limit_text}{filter_text})")

    # Build date filter clauses for JQL
    date_filter = ""
    if since_date:
        date_filter += f' AND updated >= "{since_date}"'
    if created_after:
        date_filter += f' AND created >= "{created_after}"'

    # ── Method 1: Board API ──────────────────
    if verbose:
        print(f"      → Method 1: Board API")
    try:
        board_endpoint = f"{jira_url}/rest/agile/1.0/board"
        params = {'projectKeyOrId': project_key}
        response = _get(board_endpoint, headers, auth, params=params)

        if response and response.status_code == 200:
            boards = response.json().get('values', [])
            if verbose:
                print(f"         Boards found: {len(boards)}")
            if boards:
                board_id = boards[0]['id']
                board_name = boards[0].get('name', 'Unknown')
                if verbose:
                    print(f"         Using board: {board_name} (ID: {board_id})")

                start_at = 0
                max_results = 50
                while len(all_issues) < effective_limit:
                    request_limit = max_results if unlimited else min(max_results, effective_limit - len(all_issues))
                    params = {
                        'startAt': start_at,
                        'maxResults': request_limit,
                        'fields': 'summary,description,comment,creator,created,updated',
                    }
                    if since_date:
                        params['jql'] = f'updated >= "{since_date}"'

                    resp = _get(f"{jira_url}/rest/agile/1.0/board/{board_id}/issue", headers, auth, params=params)
                    if resp and resp.status_code == 200:
                        data = resp.json()
                        issues = data.get('issues', [])
                        total = data.get('total', 0)
                        if not issues:
                            break
                        all_issues.extend(issues)
                        if verbose:
                            progress = f"{len(all_issues)}/{total}" if not unlimited else str(len(all_issues))
                            print(f"         Loaded {progress} issues")
                        if len(issues) < max_results or len(all_issues) >= total:
                            break
                        if not unlimited and len(all_issues) >= effective_limit:
                            break
                        start_at += max_results
                    else:
                        break

                if all_issues:
                    if verbose:
                        print(f"         ✅ Fetched via Board API: {len(all_issues)} issues")
                    return all_issues
    except Exception as e:
        if verbose:
            print(f"         ❌ Board API exception: {e}")

    # ── Method 2: JQL Search ────────────────
    if verbose:
        print(f"      → Method 2: JQL Search API")

    jql_variants = [
        f'project = "{project_key}"{date_filter} ORDER BY created DESC',
        f'project = {project_key}{date_filter} ORDER BY created DESC',
        f'project = "{project_key}"{date_filter}',
        f'project = {project_key}{date_filter}',
    ]

    for jql in jql_variants:
        try:
            all_issues = []
            start_at = 0
            max_results = 50
            endpoint = f"{jira_url}/rest/api/2/search"

            if verbose:
                print(f"         Trying JQL: {jql}")

            while len(all_issues) < effective_limit:
                request_limit = max_results if unlimited else min(max_results, effective_limit - len(all_issues))
                params = {
                    'jql': jql,
                    'startAt': start_at,
                    'maxResults': request_limit,
                    'fields': 'summary,description,comment,creator,created,updated',
                }
                response = _get(endpoint, headers, auth, params=params)
                if response and response.status_code == 200:
                    data = response.json()
                    total = data.get('total', 0)
                    issues = data.get('issues', [])
                    if not issues:
                        break
                    all_issues.extend(issues)
                    if verbose and start_at == 0:
                        print(f"         Total issues in project: {total}")
                    if len(issues) < max_results or len(all_issues) >= total:
                        break
                    if not unlimited and len(all_issues) >= effective_limit:
                        break
                    start_at += max_results
                else:
                    if verbose and response:
                        error_msg = response.text[:150]
                        print(f"         ⚠️  Status {response.status_code}: {error_msg}")
                    break

            if all_issues:
                if verbose:
                    print(f"         ✅ Fetched via JQL: {len(all_issues)} issues")
                return all_issues
        except Exception as e:
            if verbose:
                print(f"         ❌ Exception: {e}")
            continue

    # ── Method 3: Direct project issues access ──
    if verbose:
        print(f"      → Method 3: Project Issues API")
    try:
        response = _get(f"{jira_url}/rest/api/2/project/{project_key}", headers, auth)
        if response and response.status_code == 200:
            params = {
                'jql': f'key ~ "{project_key}-*"{date_filter}',
                'maxResults': max_results if not unlimited else 50,
                'fields': 'summary,description,comment,creator,created,updated',
            }
            response = _get(f"{jira_url}/rest/api/2/search", headers, auth, params=params)
            if response and response.status_code == 200:
                issues = response.json().get('issues', [])
                if issues:
                    if verbose:
                        print(f"         ✅ Fetched via key search: {len(issues)} issues")
                    return issues
    except Exception as e:
        if verbose:
            print(f"         ❌ Exception: {e}")

    if verbose:
        print(f"      ❌ All methods returned no results")
    return []


def scan_issue_for_secrets(issue, patterns, jira_url, email=None, api_token=None,
                           scan_attachments=False, max_attachment_size=None):
    """Scan a single issue for secrets."""
    findings = []
    issue_key = issue.get('key', 'UNKNOWN')
    issue_url = f"{jira_url}/browse/{issue_key}"
    fields = issue.get('fields', {})

    creator = fields.get('creator', {})
    author = creator.get('displayName', 'Unknown') if creator else 'Unknown'
    author_email = creator.get('emailAddress', 'N/A') if creator else 'N/A'
    created = fields.get('created', 'N/A')
    summary = fields.get('summary', '')

    texts_to_scan = []
    if summary:
        texts_to_scan.append(('Summary', summary))

    description = fields.get('description')
    if description:
        desc_text = extract_text_from_adf(description) if isinstance(description, dict) else str(description)
        if desc_text:
            texts_to_scan.append(('Description', desc_text))

    comments = fields.get('comment', {}).get('comments', [])
    for idx, comment in enumerate(comments):
        comment_body = comment.get('body')
        if comment_body:
            comment_text = extract_text_from_adf(comment_body) if isinstance(comment_body, dict) else str(comment_body)
            if comment_text:
                texts_to_scan.append((f'Comment {idx+1}', comment_text))

    if scan_attachments and email and api_token:
        for attachment in get_issue_attachments(email, api_token, jira_url, issue_key):
            att_name = attachment.get("filename", "unknown")
            text, ext = extract_text_from_attachment(attachment, email, api_token, max_size_bytes=max_attachment_size)
            if text:
                texts_to_scan.append((f'Attachment: {att_name}', text))

    for location, text in texts_to_scan:
        for secret in scan_text_for_secrets(text, patterns):
            findings.append({
                'project_key': issue_key.split('-')[0],
                'issue_key': issue_key,
                'issue_url': issue_url,
                'summary': summary,
                'author': author,
                'author_email': author_email,
                'created': created,
                'location': location,
                'secret_type': secret['secret_type'],
                'secret_value': secret['secret_value'],
                'context': secret['context'],
            })

    return findings


def extract_text_from_adf(adf_content):
    """Extract plain text from Atlassian Document Format (ADF)."""
    if not isinstance(adf_content, dict):
        return str(adf_content)

    text_parts = []

    def extract_recursive(node):
        if isinstance(node, dict):
            if 'text' in node:
                text_parts.append(node['text'])
            if 'content' in node and isinstance(node['content'], list):
                for child in node['content']:
                    extract_recursive(child)
        elif isinstance(node, list):
            for item in node:
                extract_recursive(item)

    extract_recursive(adf_content)
    return ' '.join(text_parts)


# ─────────────────────────────────────────────
# HTML report (#10)
# ─────────────────────────────────────────────
def create_html_report(findings: List[Dict], filename: str) -> str:
    """
    Create a self-contained interactive HTML report with search,
    filtering by secret type/project, and sortable columns.
    """
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    total = len(findings)
    projects = sorted(set(f['project_key'] for f in findings))
    secret_types = sorted(set(f['secret_type'] for f in findings))

    def esc(s):
        return (str(s)
                .replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;'))

    rows_html = ""
    for f in findings:
        rows_html += f"""
        <tr data-project="{esc(f['project_key'])}" data-type="{esc(f['secret_type'])}">
          <td>{esc(f['project_key'])}</td>
          <td><a href="{esc(f['issue_url'])}" target="_blank">{esc(f['issue_key'])}</a></td>
          <td>{esc(f['summary'])}</td>
          <td>{esc(f['author'])}</td>
          <td>{esc(f['created'][:10] if f['created'] != 'N/A' else 'N/A')}</td>
          <td>{esc(f['location'])}</td>
          <td><span class="badge">{esc(f['secret_type'])}</span></td>
          <td class="secret-val" title="Click to reveal value and context" onclick="toggleSecret(this)" data-val="{esc(f['secret_value'])}" data-ctx="{esc(f['context'])}">••••••••</td>
          <td class="context hidden-ctx">••••••••••••••••••••</td>
        </tr>"""

    project_options = "".join(f'<option value="{esc(p)}">{esc(p)}</option>' for p in projects)
    type_options = "".join(f'<option value="{esc(t)}">{esc(t)}</option>' for t in secret_types)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Jira Secrets Scanner Report — {timestamp}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #0f1117; color: #e0e0e0; padding: 24px; }}
  h1 {{ color: #ff4d6d; margin-bottom: 4px; font-size: 1.6rem; }}
  .subtitle {{ color: #888; font-size: .85rem; margin-bottom: 24px; }}
  .stats {{ display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }}
  .stat-card {{ background: #1a1d27; border: 1px solid #2a2d3a;
               border-radius: 8px; padding: 14px 20px; min-width: 140px; }}
  .stat-card .num {{ font-size: 2rem; font-weight: 700; color: #ff4d6d; }}
  .stat-card .lbl {{ font-size: .75rem; color: #888; text-transform: uppercase; letter-spacing: .05em; }}
  .filters {{ display: flex; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
  .filters input, .filters select {{
    background: #1a1d27; border: 1px solid #2a2d3a; color: #e0e0e0;
    padding: 8px 12px; border-radius: 6px; font-size: .85rem; outline: none; }}
  .filters input {{ flex: 1; min-width: 200px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: .82rem; }}
  thead th {{ background: #1a1d27; color: #ff4d6d; padding: 10px 8px;
              text-align: left; border-bottom: 2px solid #ff4d6d;
              white-space: nowrap; cursor: pointer; user-select: none; }}
  thead th:hover {{ background: #22263a; }}
  tbody tr {{ border-bottom: 1px solid #1e2130; transition: background .15s; }}
  tbody tr:hover {{ background: #1a1d27; }}
  td {{ padding: 8px; vertical-align: top; max-width: 260px;
        overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  td.context {{ font-family: monospace; font-size: .78rem; color: #aaa; }}
  td.hidden-ctx {{ color: #444; letter-spacing: .05em; }}
  a {{ color: #58a6ff; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .badge {{ background: #2a1a2e; color: #c084fc; border: 1px solid #6d28d9;
            padding: 2px 7px; border-radius: 4px; font-size: .75rem; white-space: nowrap; }}
  .secret-val {{ font-family: monospace; color: #ff6b6b; cursor: pointer;
                 font-size: .78rem; }}
  .secret-val.revealed {{ color: #f97316; }}
  .no-results {{ text-align: center; padding: 40px; color: #555; }}
  .sort-icon {{ margin-left: 4px; opacity: .4; }}
  .sort-icon.active {{ opacity: 1; }}
</style>
</head>
<body>
<h1>🔐 Jira Secrets Scanner</h1>
<p class="subtitle">Report generated: {timestamp}</p>

<div class="stats">
  <div class="stat-card"><div class="num">{total}</div><div class="lbl">Secrets Found</div></div>
  <div class="stat-card"><div class="num">{len(projects)}</div><div class="lbl">Projects Affected</div></div>
  <div class="stat-card"><div class="num">{len(set(f['issue_key'] for f in findings))}</div><div class="lbl">Issues Affected</div></div>
  <div class="stat-card"><div class="num">{len(secret_types)}</div><div class="lbl">Secret Types</div></div>
</div>

<div class="filters">
  <input type="text" id="search" placeholder="🔍  Search across all columns…" oninput="applyFilters()">
  <select id="filterProject" onchange="applyFilters()">
    <option value="">All projects</option>{project_options}
  </select>
  <select id="filterType" onchange="applyFilters()">
    <option value="">All secret types</option>{type_options}
  </select>
</div>

<table id="findings-table">
<thead>
  <tr>
    <th onclick="sortTable(0)">Project<span class="sort-icon" id="si0">▲</span></th>
    <th onclick="sortTable(1)">Issue<span class="sort-icon" id="si1">▲</span></th>
    <th onclick="sortTable(2)">Summary<span class="sort-icon" id="si2">▲</span></th>
    <th onclick="sortTable(3)">Author<span class="sort-icon" id="si3">▲</span></th>
    <th onclick="sortTable(4)">Created<span class="sort-icon" id="si4">▲</span></th>
    <th onclick="sortTable(5)">Location<span class="sort-icon" id="si5">▲</span></th>
    <th onclick="sortTable(6)">Secret Type<span class="sort-icon" id="si6">▲</span></th>
    <th>Value (click)</th>
    <th>Context</th>
  </tr>
</thead>
<tbody id="table-body">
{rows_html}
</tbody>
</table>
<p id="no-results" class="no-results" style="display:none">No matching findings.</p>

<script>
let sortCol = -1, sortAsc = true;

function applyFilters() {{
  const q = document.getElementById('search').value.toLowerCase();
  const proj = document.getElementById('filterProject').value;
  const type = document.getElementById('filterType').value;
  let visible = 0;
  document.querySelectorAll('#table-body tr').forEach(row => {{
    const text = row.innerText.toLowerCase();
    const matchQ = !q || text.includes(q);
    const matchP = !proj || row.dataset.project === proj;
    const matchT = !type || row.dataset.type === type;
    const show = matchQ && matchP && matchT;
    row.style.display = show ? '' : 'none';
    if (show) visible++;
  }});
  document.getElementById('no-results').style.display = visible === 0 ? 'block' : 'none';
}}

function sortTable(col) {{
  const tbody = document.getElementById('table-body');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  if (sortCol === col) sortAsc = !sortAsc; else {{ sortCol = col; sortAsc = true; }}
  document.querySelectorAll('.sort-icon').forEach((el, i) => {{
    el.textContent = i === col ? (sortAsc ? '▲' : '▼') : '▲';
    el.classList.toggle('active', i === col);
  }});
  rows.sort((a, b) => {{
    const av = a.cells[col]?.innerText.trim() || '';
    const bv = b.cells[col]?.innerText.trim() || '';
    return sortAsc ? av.localeCompare(bv) : bv.localeCompare(av);
  }});
  rows.forEach(r => tbody.appendChild(r));
}}

function toggleSecret(cell) {{
  const ctx = cell.nextElementSibling;
  const revealed = cell.classList.contains('revealed');
  if (revealed) {{
    cell.textContent = '••••••••';
    cell.classList.remove('revealed');
    cell.style.whiteSpace = 'nowrap';
    cell.style.maxWidth = '';
    ctx.textContent = '••••••••••••••••••••';
    ctx.classList.add('hidden-ctx');
    ctx.style.whiteSpace = 'nowrap';
    ctx.style.maxWidth = '';
  }} else {{
    cell.textContent = cell.dataset.val;
    cell.classList.add('revealed');
    cell.style.whiteSpace = 'normal';
    cell.style.maxWidth = '340px';
    ctx.textContent = cell.dataset.ctx;
    ctx.classList.remove('hidden-ctx');
    ctx.style.whiteSpace = 'normal';
    ctx.style.maxWidth = '400px';
  }}
}}
</script>
</body>
</html>"""

    with open(filename, 'w', encoding='utf-8') as fh:
        fh.write(html)
    return filename


# ─────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────
def create_secrets_report(findings, filename="jira_secrets_report.xlsx"):
    """Create a formatted Excel report with all findings."""
    wb = Workbook()

    sheet_secrets = wb.active
    sheet_secrets.title = "Found Secrets"

    headers = [
        'Project', 'Issue Key', 'Issue URL', 'Summary', 'Author', 'Author Email',
        'Created', 'Location', 'Secret Type', 'Secret Value', 'Context'
    ]

    header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = sheet_secrets.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_num, finding in enumerate(findings, 2):
        sheet_secrets.cell(row=row_num, column=1).value = finding['project_key']
        sheet_secrets.cell(row=row_num, column=2).value = finding['issue_key']

        url_cell = sheet_secrets.cell(row=row_num, column=3)
        url_cell.value = finding['issue_url']
        url_cell.hyperlink = finding['issue_url']
        url_cell.font = Font(color='0563C1', underline='single')

        sheet_secrets.cell(row=row_num, column=4).value = finding['summary']
        sheet_secrets.cell(row=row_num, column=5).value = finding['author']
        sheet_secrets.cell(row=row_num, column=6).value = finding['author_email']
        sheet_secrets.cell(row=row_num, column=7).value = finding['created']
        sheet_secrets.cell(row=row_num, column=8).value = finding['location']
        sheet_secrets.cell(row=row_num, column=9).value = finding['secret_type']

        secret_cell = sheet_secrets.cell(row=row_num, column=10)
        raw_value = str(finding['secret_value'])
        secret_cell.value = "'" + raw_value if raw_value.startswith(('=', '+', '-', '@')) else raw_value
        secret_cell.font = Font(color='DC143C', bold=True)

        sheet_secrets.cell(row=row_num, column=11).value = finding['context']

        for col_num in range(1, len(headers) + 1):
            sheet_secrets.cell(row=row_num, column=col_num).border = border

    sheet_secrets.column_dimensions['A'].width = 12
    sheet_secrets.column_dimensions['B'].width = 15
    sheet_secrets.column_dimensions['C'].width = 45
    sheet_secrets.column_dimensions['D'].width = 40
    sheet_secrets.column_dimensions['E'].width = 20
    sheet_secrets.column_dimensions['F'].width = 25
    sheet_secrets.column_dimensions['G'].width = 20
    sheet_secrets.column_dimensions['H'].width = 15
    sheet_secrets.column_dimensions['I'].width = 30
    sheet_secrets.column_dimensions['J'].width = 50
    sheet_secrets.column_dimensions['K'].width = 60

    for row in range(2, len(findings) + 2):
        for col in [4, 10, 11]:
            sheet_secrets.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    sheet_secrets.freeze_panes = 'A2'

    sheet_stats = wb.create_sheet("Statistics")
    sheet_stats['A1'] = 'Secret Scanning Statistics'
    sheet_stats['A1'].font = Font(bold=True, size=14)

    stats_data = [
        ['Total Secrets Found:', len(findings)],
        ['Unique Secret Types:', len(set(f['secret_type'] for f in findings))],
        ['Affected Projects:', len(set(f['project_key'] for f in findings))],
        ['Affected Issues:', len(set(f['issue_key'] for f in findings))],
    ]

    for idx, (label, value) in enumerate(stats_data, 3):
        sheet_stats.cell(row=idx, column=1).value = label
        sheet_stats.cell(row=idx, column=1).font = Font(bold=True)
        sheet_stats.cell(row=idx, column=2).value = value

    sheet_stats.cell(row=len(stats_data) + 5, column=1).value = 'Secrets by Type:'
    sheet_stats.cell(row=len(stats_data) + 5, column=1).font = Font(bold=True, size=12)

    secret_types = {}
    for finding in findings:
        secret_types[finding['secret_type']] = secret_types.get(finding['secret_type'], 0) + 1

    row_offset = len(stats_data) + 6
    for idx, (secret_type, count) in enumerate(sorted(secret_types.items(), key=lambda x: x[1], reverse=True)):
        sheet_stats.cell(row=row_offset + idx, column=1).value = secret_type
        sheet_stats.cell(row=row_offset + idx, column=2).value = count

    sheet_stats.column_dimensions['A'].width = 35
    sheet_stats.column_dimensions['B'].width = 15

    wb.save(filename)
    return filename


def export_findings_to_json(findings: List[Dict], filename: str, scan_stats: Dict = None) -> str:
    """Export findings to a JSON file."""
    output = {
        "generated_at": datetime.now().isoformat(),
        "scan_stats": scan_stats or {},
        "total_findings": len(findings),
        "findings": findings,
    }
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)
    return filename


# ─────────────────────────────────────────────
# Email — multiple recipients (#new)
# ─────────────────────────────────────────────
def send_email_report(report_filename, findings, scan_stats, email_config):
    """
    Send email report via AWS SES.

    email_config['recipient'] can be a single address or a comma-separated
    string of multiple addresses.
    """
    if not AWS_SES_AVAILABLE:
        print("❌ AWS SES not available. Install: pip install boto3")
        return False

    sender = email_config.get('sender')
    recipient_raw = email_config.get('recipient', '')
    aws_region = email_config.get('aws_region', 'eu-central-1')

    if not sender or not recipient_raw:
        print("❌ Email sender and recipient are required")
        return False

    # Parse multiple recipients
    recipients = [r.strip() for r in recipient_raw.split(',') if r.strip()]
    if not recipients:
        print("❌ No valid recipient addresses found")
        return False

    try:
        ses_client = boto3.client("ses", region_name=aws_region)

        total_secrets = len(findings)
        subject = (
            f"CRITICAL: Jira Secrets Scanner - {total_secrets} Secret{'s' if total_secrets != 1 else ''} Found"
            if total_secrets > 0
            else "INFO: Jira Secrets Scanner - No Secrets Found"
        )

        body_text = generate_email_body(findings, scan_stats)

        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)   # display all in To: header
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "plain"))

        if os.path.exists(report_filename):
            with open(report_filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment; filename=jira_secrets.xlsx")
            msg.attach(part)

        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=recipients,   # SES requires a list
            RawMessage={"Data": msg.as_string()},
        )

        print(f"\n✅ Email sent successfully!")
        print(f"   MessageId : {response['MessageId']}")
        print(f"   From      : {sender}")
        print(f"   To        : {', '.join(recipients)}")
        print(f"   Subject   : {subject}")
        return True

    except Exception as e:
        print(f"\n❌ Error sending email: {e}")
        return False


def generate_email_body(findings, scan_stats):
    """Generate email body text with scan results."""
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    body = f"""Jira Secrets Scanner Report - {timestamp}

Summary Statistics:
* Total Secrets Found: {len(findings)}
* Affected Projects: {len(set(f['project_key'] for f in findings)) if findings else 0}
* Affected Issues: {len(set(f['issue_key'] for f in findings)) if findings else 0}

"""
    if findings:
        body += """ACTION REQUIRED:
1. Review the attached report immediately
2. Rotate/revoke exposed credentials
3. Implement proper secrets management

The detailed report is attached as XLSX file.

---
This is an automated report generated by Jira Secrets Scanner."""
    else:
        body += """RESULT:
No secrets detected in scanned Jira issues.
This is a good security posture!

---
This is an automated report generated by Jira Secrets Scanner."""
    return body


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
def parse_arguments():
    parser = argparse.ArgumentParser(
        description='Scan Jira projects and issues for exposed secrets.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan all projects
  python jira_scanner.py --env --scan-secrets

  # Scan specific projects
  python jira_scanner.py -e user@company.com -t TOKEN -u URL --scan-secrets --projects PROJ1,PROJ2

  # Limit issues per project
  python jira_scanner.py --env --scan-secrets --max-issues 50

  # Incremental scan (only issues updated since last run)
  python jira_scanner.py --env --scan-secrets --incremental

  # Incremental scan with custom state file
  python jira_scanner.py --env --scan-secrets --incremental --state-file /tmp/my_state.json

  # Use custom patterns file
  python jira_scanner.py --env --scan-secrets --patterns custom_patterns.txt

  # Use TruffleHog YAML with all detectors
  python jira_scanner.py --env --scan-secrets -tp detectors.yaml

  # Only include AWS-related TruffleHog detectors
  python jira_scanner.py --env --scan-secrets -tp detectors.yaml -tk aws

  # Include aws and api detectors, but exclude gateway ones
  python jira_scanner.py --env --scan-secrets -tp detectors.yaml -tk aws,api -tek gateway,arn

  # Export findings as both XLSX and JSON
  python jira_scanner.py --env --scan-secrets --json -o jira_secrets

  # Also generate HTML report
  python jira_scanner.py --env --scan-secrets --html -o jira_secrets

  # Skip known false positives
  python jira_scanner.py --env --scan-secrets --ignore-file .jira_scanner_ignore

  # Send to multiple recipients
  python jira_scanner.py --env --scan-secrets --email-sender a@x.com --email-recipient "b@x.com,c@x.com"

  # Parallel scanning (faster for large instances)
  python jira_scanner.py --env --scan-secrets --workers 20
        """
    )

    parser.add_argument('-e', '--email', type=str, help='Atlassian email')
    parser.add_argument('-t', '--token', type=str, help='Atlassian API token')
    parser.add_argument('-u', '--url', type=str, help='Jira instance URL')
    parser.add_argument('-o', '--output', type=str, help='Output filename base (extensions added automatically)')
    parser.add_argument('--json', action='store_true', help='Export findings to JSON')
    parser.add_argument('--html', action='store_true', help='Export findings to interactive HTML report')  # #10
    parser.add_argument('--env', action='store_true', help='Load credentials from .env file')
    parser.add_argument('--env-file', type=str, default='.env', help='Path to .env file')
    parser.add_argument('--scan-secrets', action='store_true', help='Enable secret scanning')
    parser.add_argument('--patterns', type=str, default='secret_patterns.txt',
                        help='Secret patterns file (Name:::Regex:::GroupIndex format)')
    parser.add_argument('--trufflehog-patterns', '-tp', type=str, default=None,
                        help='Path to a TruffleHog v3 YAML file with detectors.')
    parser.add_argument('--trufflehog-keywords', '-tk', type=str, default=None,
                        help='Include only TruffleHog detectors matching these keywords (comma-separated).')
    parser.add_argument('--trufflehog-exclude-keywords', '-tek', type=str, default=None,
                        help='Exclude TruffleHog detectors matching these keywords (comma-separated).')
    parser.add_argument('--projects', type=str, help='Comma-separated project keys to scan (default: all)')
    parser.add_argument('--max-issues', type=int, default=0,
                        help='Max issues per project; 0 = unlimited (default: 0)')
    parser.add_argument('-q', '--quiet', action='store_true', help='Suppress per-issue output')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose debug output')
    parser.add_argument('--no-duplicates', '-nd', action='store_true',
                        help='Deduplicate findings by (issue_key, secret_value)')

    # Incremental scan (#17)
    parser.add_argument('--incremental', action='store_true',
                        help='Only scan issues updated since the last run (state stored in --state-file)')
    parser.add_argument('--state-file', type=str, default=_STATE_FILE,
                        help=f'Path to incremental scan state file (default: {_STATE_FILE})')

    # Date filters
    parser.add_argument('--since-days', type=int, default=None,
                        help='Only scan issues updated in the last N days (e.g. --since-days 7)')
    parser.add_argument('--since-date', type=str, default=None,
                        help='Only scan issues updated on or after this date (YYYY-MM-DD, e.g. --since-date 2026-01-01)')
    parser.add_argument('--created-after', type=str, default=None,
                        help='Only scan issues created on or after this date (YYYY-MM-DD, e.g. --created-after 2026-01-01)')

    # Whitelist (#18)
    parser.add_argument('--ignore-file', type=str, default=None,
                        help='Path to ignore/whitelist file (format: ISSUE-KEY:SecretType:SecretValue, '
                             'use * as wildcard for issue key). Default: not used.')

    # Parallel workers (#4)
    parser.add_argument('--workers', type=int, default=1,
                        help='Number of parallel threads for scanning issues (default: 1, recommended: 5-20)')

    # Attachment scanning
    parser.add_argument('--scan-attachments', action='store_true',
                        help='Scan issue attachments (txt, json, py, pdf, docx, images via OCR)')
    parser.add_argument('--max-attachment-size', type=str, default=None,
                        help='Max attachment size to scan, e.g. 2mb, 500kb (default: no limit)')

    # Email
    parser.add_argument('--email-sender', type=str, help='Sender email address (must be verified in SES)')
    parser.add_argument('--email-recipient', type=str,
                        help='Recipient email address(es) — comma-separated for multiple recipients')
    parser.add_argument('--aws-region', type=str, default='eu-central-1',
                        help='AWS region for SES (default: eu-central-1)')

    return parser.parse_args()


def parse_size(size_str):
    """Parse size string like '2mb' or '500kb' to bytes. Returns None if input is None."""
    if not size_str:
        return None
    size_str = size_str.lower().strip()
    match = re.match(r"(\d+)(kb|mb|gb)?", size_str)
    if not match:
        return None
    num = int(match.group(1))
    unit = match.group(2) or "b"
    multipliers = {"b": 1, "kb": 1024, "mb": 1024 ** 2, "gb": 1024 ** 3}
    return num * multipliers.get(unit, 1)


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main():
    global _partial_findings, _interrupted

    args = parse_arguments()

    print("🔍 Jira Secret Scanner")
    print("=" * 80)

    env_vars = {}
    if args.env:
        env_vars = load_env_file(args.env_file)
        if env_vars:
            print(f"✅ Loaded config from {args.env_file}\n")

    email = args.email or env_vars.get('JIRA_EMAIL') or input("Enter your Atlassian email: ").strip()
    api_token = args.token or env_vars.get('JIRA_TOKEN') or input("Enter your API token: ").strip()
    jira_url = args.url or env_vars.get('JIRA_URL') or input("Enter your Jira URL: ").strip()

    if not email or not api_token or not jira_url:
        print("❌ Email, token, and URL are required.")
        sys.exit(1)

    jira_url = normalize_jira_url(jira_url)

    print("\n🔄 Fetching projects…")
    projects = get_jira_projects(email, api_token, jira_url)
    if not projects:
        print("❌ Failed to fetch projects.")
        sys.exit(1)
    print(f"✅ Projects found: {len(projects)}\n")

    if args.projects:
        selected_keys = [k.strip().upper() for k in args.projects.split(',')]
        projects = [p for p in projects if p.get('key', '').upper() in selected_keys]
        print(f"🎯 Projects selected for scanning: {len(projects)}")

    if not args.scan_secrets:
        print("ℹ️  Use --scan-secrets flag to enable secret scanning.")
        return

    print("\n🔐 Starting secret scan…\n")

    # ── Attachment scanning setup ──
    scan_attachments = args.scan_attachments
    max_attachment_size = parse_size(args.max_attachment_size)
    if scan_attachments:
        print("🖼️  OCR enabled" if OCR_AVAILABLE else "⚠️  OCR unavailable")
        print("📄 PDF support enabled" if PDF_AVAILABLE else "⚠️  PDF support unavailable")
        print("📝 DOCX support enabled" if DOCX_AVAILABLE else "⚠️  DOCX support unavailable")
        print(f"📦 Max attachment size: {args.max_attachment_size or 'unlimited'}\n")

    # ── Pattern validation ──
    if args.trufflehog_keywords and not args.trufflehog_patterns:
        print("❌ -tk / --trufflehog-keywords requires -tp / --trufflehog-patterns.")
        sys.exit(1)
    if args.trufflehog_exclude_keywords and not args.trufflehog_patterns:
        print("❌ -tek / --trufflehog-exclude-keywords requires -tp / --trufflehog-patterns.")
        sys.exit(1)

    # ── Load patterns (compiled, #6) ──
    if args.trufflehog_patterns:
        if not Path(args.trufflehog_patterns).exists():
            print(f"❌ TruffleHog patterns file '{args.trufflehog_patterns}' not found.")
            sys.exit(1)
        include_kw = [k.strip() for k in args.trufflehog_keywords.split(',')] if args.trufflehog_keywords else None
        exclude_kw = [k.strip() for k in args.trufflehog_exclude_keywords.split(',')] if args.trufflehog_exclude_keywords else None
        patterns, th_stats = load_trufflehog_patterns(args.trufflehog_patterns, include_kw, exclude_kw)
        print(f"✅ TruffleHog patterns file: {args.trufflehog_patterns}")
        print(f"   Total detectors : {th_stats['total']}")
        if include_kw:
            print(f"   Skipped (-tk)   : {th_stats['skipped_include']}")
        if exclude_kw:
            print(f"   Skipped (-tek)  : {th_stats['skipped_exclude']}")
        print(f"   Active          : {th_stats['loaded']}\n")
    else:
        patterns = load_secret_patterns(args.patterns)
        print(f"✅ Patterns loaded ('{args.patterns}'): {len(patterns)}\n")

    # ── Ignore list (#18) ──
    ignore_set = set()
    if args.ignore_file:
        ignore_set = load_ignore_list(args.ignore_file)
    elif Path('.jira_scanner_ignore').exists():
        # Auto-load default ignore file if present
        ignore_set = load_ignore_list('.jira_scanner_ignore')

    # ── Incremental scan state (#17) ──
    scan_state = {}
    if args.incremental:
        scan_state = load_scan_state(args.state_file)
        if scan_state:
            print(f"🔄 Incremental mode — loaded state from '{args.state_file}'")
        else:
            print(f"🔄 Incremental mode — no previous state found, performing full scan")

    # ── Date filter flags ──
    # --since-days takes priority over --since-date if both provided
    global_since_date = None
    if args.since_days is not None:
        from datetime import timedelta
        global_since_date = (datetime.now() - timedelta(days=args.since_days)).strftime('%Y-%m-%d')
        print(f"📅 Date filter: issues updated in the last {args.since_days} day(s) (since {global_since_date})")
    elif args.since_date is not None:
        global_since_date = args.since_date
        print(f"📅 Date filter: issues updated on or after {global_since_date}")

    global_created_after = None
    if args.created_after is not None:
        global_created_after = args.created_after
        print(f"📅 Date filter: issues created on or after {global_created_after}")

    if global_since_date or global_created_after:
        print()

    # ── Workers info ──
    if args.workers > 1:
        print(f"⚡ Parallel scanning enabled: {args.workers} workers\n")

    all_findings: List[Dict] = []
    total_issues_scanned = 0
    new_scan_state = {}

    for project in projects:
        if _interrupted:
            break

        project_key = project.get('key', 'UNKNOWN')
        project_name = project.get('name', 'Unknown')
        print(f"📊 Scanning project: {project_key} — {project_name}")

        # Determine updated-since filter: incremental state takes priority,
        # then --since-days / --since-date
        since_date = global_since_date
        if args.incremental and project_key in scan_state:
            incremental_date = scan_state[project_key].get('last_scan')
            if incremental_date:
                since_date = incremental_date
                print(f"   ⏩ Incremental: only issues updated since {since_date}")

        created_after = global_created_after

        issues = get_project_issues(
            email, api_token, jira_url, project_key,
            args.max_issues, args.verbose,
            since_date=since_date, created_after=created_after
        )
        print(f"   Issues fetched: {len(issues)}")
        total_issues_scanned += len(issues)

        # Record scan time for this project (before scanning, to avoid missing
        # issues updated mid-scan on next run)
        new_scan_state[project_key] = {'last_scan': datetime.now().strftime('%Y-%m-%d')}

        # ── Parallel or sequential scanning (#4) ──
        project_findings: List[Dict] = []

        def _scan_one(issue):
            return scan_issue_for_secrets(
                issue, patterns, jira_url,
                email=email, api_token=api_token,
                scan_attachments=scan_attachments,
                max_attachment_size=max_attachment_size,
            )

        # Only show tqdm bar when running in a real TTY (not Docker/pipe)
        use_tqdm = TQDM_AVAILABLE and not args.quiet and sys.stderr.isatty()

        if args.workers > 1:
            # Parallel path (#4)
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                future_map = {executor.submit(_scan_one, issue): issue for issue in issues}
                if use_tqdm:
                    futures_iter = tqdm(as_completed(future_map), total=len(future_map),
                                        desc=f"  {project_key}", unit="issue", leave=False)
                else:
                    futures_iter = as_completed(future_map)
                for future in futures_iter:
                    if _interrupted:
                        executor.shutdown(wait=False, cancel_futures=True)
                        break
                    try:
                        findings = future.result()
                        project_findings.extend(findings)
                    except Exception as exc:
                        issue_key = future_map[future].get('key', '?')
                        tqdm.write(f"   ⚠️  Error scanning {issue_key}: {exc}") if use_tqdm else print(f"   ⚠️  Error scanning {issue_key}: {exc}")
        else:
            # Sequential path with optional tqdm (#15)
            if use_tqdm:
                issue_iter = tqdm(issues, desc=f"  {project_key}", unit="issue", leave=False)
            else:
                issue_iter = issues

            for issue in issue_iter:
                if _interrupted:
                    break
                findings = _scan_one(issue)
                project_findings.extend(findings)

        # ── Apply ignore list (#18) ──
        before_ignore = len(project_findings)
        project_findings = [f for f in project_findings if not is_ignored(f, ignore_set)]
        ignored_count = before_ignore - len(project_findings)
        def _log(msg):
            tqdm.write(msg) if use_tqdm else print(msg)

        if ignored_count and not args.quiet:
            _log(f"   🚫 Ignored (whitelist): {ignored_count}")

        # ── Deduplication ──
        if args.no_duplicates:
            seen = {(f['issue_key'], f['secret_value']) for f in all_findings}
            deduped = []
            for f in project_findings:
                key = (f['issue_key'], f['secret_value'])
                if key not in seen:
                    seen.add(key)
                    deduped.append(f)
            removed = len(project_findings) - len(deduped)
            if removed and not args.quiet:
                _log(f"   🔁 Deduplicated: {removed}")
            project_findings = deduped

        if project_findings and not args.quiet:
            _log(f"   ⚠️  Secrets found: {len(project_findings)}")

        all_findings.extend(project_findings)
        _partial_findings = all_findings  # keep global ref for Ctrl+C handler (#3)
        print()

    # ── Save incremental state (#17) ──
    if args.incremental:
        # Merge new state with old (keep projects we didn't scan this time)
        merged_state = {**scan_state, **new_scan_state}
        save_scan_state(merged_state, args.state_file)
        print(f"💾 Scan state saved to '{args.state_file}'")

    # ── Summary ──
    dedup_note = " (duplicates removed)" if args.no_duplicates else ""
    interrupted_note = " ⚠️  SCAN INTERRUPTED — partial results" if _interrupted else ""
    print(f"\n📊 Total secrets found: {len(all_findings)}{dedup_note}{interrupted_note}\n")

    scan_stats = {
        'projects_scanned': len(projects),
        'issues_scanned': total_issues_scanned,
        'interrupted': _interrupted,
    }

    report_filename = None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = (
        re.sub(r'\.(xlsx|json|html)$', '', args.output, flags=re.IGNORECASE)
        if args.output
        else f"jira_secrets_report_{timestamp}"
    )

    if all_findings:
        print("📝 Generating XLSX report…")
        xlsx_filename = base_name + ".xlsx"
        create_secrets_report(all_findings, xlsx_filename)
        print(f"✅ XLSX report saved: {xlsx_filename}")
        report_filename = xlsx_filename

        if args.json:
            json_filename = base_name + ".json"
            print("📝 Generating JSON report…")
            export_findings_to_json(all_findings, json_filename, scan_stats)
            print(f"✅ JSON report saved: {json_filename}")

        if args.html:
            html_filename = base_name + ".html"
            print("📝 Generating HTML report…")
            create_html_report(all_findings, html_filename)
            print(f"✅ HTML report saved: {html_filename}")

        print(f"\n📈 Summary:")
        print(f"   Affected projects : {len(set(f['project_key'] for f in all_findings))}")
        print(f"   Affected issues   : {len(set(f['issue_key'] for f in all_findings))}")
        print(f"   Secret types      : {len(set(f['secret_type'] for f in all_findings))}")
    else:
        print("✅ No secrets found.")
        if args.json:
            json_filename = base_name + ".json"
            export_findings_to_json([], json_filename, scan_stats)
            print(f"✅ JSON report saved: {json_filename}")

    # ── Email ──
    if args.email_sender and args.email_recipient:
        print("\n📧 Sending email report…")
        email_config = {
            'sender': args.email_sender,
            'recipient': args.email_recipient,
            'aws_region': args.aws_region,
        }

        # Create temp XLSX if there were no findings (so we always attach something)
        temp_report = False
        if not report_filename:
            report_filename = f"jira_secrets_report_temp_{timestamp}.xlsx"
            create_secrets_report(all_findings, report_filename)
            temp_report = True

        send_email_report(report_filename, all_findings, scan_stats, email_config)

        # Clean up temp file (#2)
        if temp_report and os.path.exists(report_filename):
            os.remove(report_filename)


if __name__ == "__main__":
    main()
