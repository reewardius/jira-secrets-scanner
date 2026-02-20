#!/usr/bin/env python3
"""
Скрипт для получения списка доступных Jira проектов (spaces)
и сканирования тикетов на наличие секретов
Использует Atlassian API token и email для авторизации
Выводит результаты в Excel файл с подробной информацией
"""

import requests
from requests.auth import HTTPBasicAuth
import json
import sys
import argparse
import os
import re
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Tuple

# AWS SES для отправки email
try:
    import boto3
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    AWS_SES_AVAILABLE = True
except ImportError:
    AWS_SES_AVAILABLE = False


def load_env_file(env_path='.env'):
    """
    Загружает переменные из .env файла
    """
    env_vars = {}
    env_file = Path(env_path)
    
    if env_file.exists():
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_vars[key.strip()] = value.strip()
    
    return env_vars


def normalize_jira_url(url):
    """
    Нормализует Jira URL, удаляя конечный слеш
    """
    return url.rstrip('/')


def load_secret_patterns(patterns_file='secret_patterns.txt'):
    """
    Загружает паттерны для поиска секретов из файла
    Формат: Name:::Regex:::GroupIndex
    """
    patterns = []
    
    if not Path(patterns_file).exists():
        print(f"⚠️  Файл с паттернами {patterns_file} не найден. Используются базовые паттерны.")
        # Базовые паттерны если файл не найден
        return [
            ('AWS Access Key ID', r'(?:^|[^A-Za-z0-9])((AKIA|ASIA|AGPA|AIDA|AROA|AIPA|ANPA|ANVA)(?!([A-Z0-9])\3{5,})[A-Z0-9]{16})(?:[^A-Za-z0-9]|$)', 1),
            ('GitHub Personal Access Token', r'(?:^|[^a-z0-9_])(ghp_[0-9a-zA-Z]{36})(?:[^a-zA-Z0-9]|$)', 1),
            ('Atlassian API Token', r'(?:^|[^A-Z])(ATATT[a-zA-Z0-9\-_]{28,})(?:[^a-zA-Z0-9\-_]|$)', 1),
        ]
    
    with open(patterns_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and ':::' in line:
                parts = line.split(':::')
                if len(parts) >= 3:
                    name = parts[0].strip()
                    regex = parts[1].strip()
                    group_index = int(parts[2].strip())
                    patterns.append((name, regex, group_index))
    
    return patterns


def scan_text_for_secrets(text: str, patterns: List[Tuple[str, str, int]]) -> List[Dict]:
    """
    Сканирует текст на наличие секретов используя предоставленные паттерны
    
    Returns:
        List of dicts with keys: secret_type, secret_value, context
    """
    findings = []
    
    for pattern_name, pattern_regex, group_index in patterns:
        try:
            matches = re.finditer(pattern_regex, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                if group_index < len(match.groups()) + 1:
                    secret_value = match.group(group_index) if group_index > 0 else match.group(0)
                    
                    # Получаем контекст (50 символов до и после)
                    start = max(0, match.start() - 50)
                    end = min(len(text), match.end() + 50)
                    context = text[start:end].replace('\n', ' ').replace('\r', '').strip()
                    
                    findings.append({
                        'secret_type': pattern_name,
                        'secret_value': secret_value,
                        'context': context
                    })
        except re.error as e:
            print(f"⚠️  Ошибка в regex паттерне '{pattern_name}': {e}")
            continue
    
    return findings


def get_jira_projects(email, api_token, jira_url):
    """
    Получает список всех доступных Jira проектов с подробной информацией
    """
    jira_url = normalize_jira_url(jira_url)
    url = f"{jira_url}/rest/api/3/project/search"
    auth = HTTPBasicAuth(email, api_token)
    params = {'expand': 'description,lead,url,insight'}
    headers = {"Accept": "application/json"}
    
    all_projects = []
    start_at = 0
    max_results = 50
    
    try:
        while True:
            params['startAt'] = start_at
            params['maxResults'] = max_results
            
            response = requests.get(url, headers=headers, auth=auth, params=params)
            
            if response.status_code == 200:
                data = response.json()
                projects = data.get('values', [])
                all_projects.extend(projects)
                
                if data.get('isLast', True):
                    break
                start_at += max_results
            elif response.status_code == 401:
                print("❌ Ошибка авторизации. Проверьте email и API token.")
                return None
            elif response.status_code == 403:
                print("❌ Доступ запрещен. Проверьте права доступа.")
                return None
            else:
                print(f"❌ Ошибка: {response.status_code}")
                print(f"Ответ: {response.text}")
                return None
        
        return all_projects
    except requests.exceptions.RequestException as e:
        print(f"❌ Ошибка при запросе: {e}")
        return None


def get_project_issues(email, api_token, jira_url, project_key, max_issues=0, verbose=False):
    """
    Получает список issues для конкретного проекта
    Пробует разные API endpoints для совместимости
    max_issues=0 означает получить все issues
    """
    jira_url = normalize_jira_url(jira_url)
    auth = HTTPBasicAuth(email, api_token)
    headers = {"Accept": "application/json"}
    
    all_issues = []
    
    # Если max_issues=0, устанавливаем очень большой лимит
    unlimited = (max_issues == 0)
    if unlimited:
        effective_limit = 999999  # Практически безлимит
    else:
        effective_limit = max_issues
    
    if verbose:
        limit_text = "все" if unlimited else str(max_issues)
        print(f"   🔍 Начинаю поиск issues для проекта {project_key} (лимит: {limit_text})")
    
    # Метод 1: Board API (для Software/Scrum/Kanban проектов)
    if verbose:
        print(f"      → Метод 1: Board API (для Software проектов)")
    
    try:
        # Ищем boards для проекта
        board_endpoint = f"{jira_url}/rest/agile/1.0/board"
        params = {'projectKeyOrId': project_key}
        response = requests.get(board_endpoint, headers=headers, auth=auth, params=params, timeout=30)
        
        if verbose:
            print(f"         Board search: {response.status_code}")
        
        if response.status_code == 200:
            boards_data = response.json()
            boards = boards_data.get('values', [])
            
            if verbose:
                print(f"         Найдено boards: {len(boards)}")
            
            if boards:
                board_id = boards[0]['id']
                board_name = boards[0].get('name', 'Unknown')
                if verbose:
                    print(f"         Использую board: {board_name} (ID: {board_id})")
                
                # Получаем issues через board с пагинацией
                start_at = 0
                max_results = 50
                
                while len(all_issues) < effective_limit:
                    issues_endpoint = f"{jira_url}/rest/agile/1.0/board/{board_id}/issue"
                    
                    # Определяем сколько issues запрашивать в этой итерации
                    if unlimited:
                        request_limit = max_results
                    else:
                        request_limit = min(max_results, effective_limit - len(all_issues))
                    
                    params = {
                        'startAt': start_at,
                        'maxResults': request_limit,
                        'fields': 'summary,description,comment,creator,created,updated'
                    }
                    
                    response = requests.get(issues_endpoint, headers=headers, auth=auth, params=params, timeout=30)
                    
                    if response.status_code == 200:
                        data = response.json()
                        issues = data.get('issues', [])
                        total = data.get('total', 0)
                        
                        if not issues:
                            break
                        
                        all_issues.extend(issues)
                        
                        if verbose:
                            progress = f"{len(all_issues)}/{total}" if not unlimited else f"{len(all_issues)}"
                            print(f"         Загружено {progress} issues")
                        
                        # Проверяем условия выхода
                        if len(issues) < max_results:  # Последняя страница
                            break
                        if len(all_issues) >= total:  # Получили все
                            break
                        if not unlimited and len(all_issues) >= effective_limit:  # Достигли лимита
                            break
                        
                        start_at += max_results
                    else:
                        if verbose:
                            print(f"         ⚠️  Ошибка пагинации: {response.status_code}")
                        break
                
                if all_issues:
                    if verbose:
                        print(f"         ✅ Получено через Board API: {len(all_issues)} issues")
                    return all_issues
                        
    except Exception as e:
        if verbose:
            print(f"         ❌ Board API исключение: {e}")
    
    # Метод 2: JQL Search (классический метод)
    if verbose:
        print(f"      → Метод 2: JQL Search API")
    
    jql_variants = [
        f'project = "{project_key}" ORDER BY created DESC',
        f'project = {project_key} ORDER BY created DESC',
        f'project = "{project_key}"',
        f'project = {project_key}',
    ]
    
    for jql in jql_variants:
        try:
            all_issues = []
            start_at = 0
            max_results = 50
            
            # Пробуем API v2 (более совместимый)
            endpoint = f"{jira_url}/rest/api/2/search"
            
            if verbose:
                print(f"         Пробую JQL: {jql}")
            
            while len(all_issues) < effective_limit:
                # Определяем сколько issues запрашивать
                if unlimited:
                    request_limit = max_results
                else:
                    request_limit = min(max_results, effective_limit - len(all_issues))
                
                params = {
                    'jql': jql,
                    'startAt': start_at,
                    'maxResults': request_limit,
                    'fields': 'summary,description,comment,creator,created,updated'
                }
                
                response = requests.get(endpoint, headers=headers, auth=auth, params=params, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    total = data.get('total', 0)
                    issues = data.get('issues', [])
                    
                    if not issues:
                        break
                    
                    all_issues.extend(issues)
                    
                    if verbose and start_at == 0:
                        print(f"         Total issues в проекте: {total}")
                    
                    # Проверяем условия выхода
                    if len(issues) < max_results:  # Последняя страница
                        break
                    if len(all_issues) >= total:  # Получили все
                        break
                    if not unlimited and len(all_issues) >= effective_limit:  # Достигли лимита
                        break
                    
                    start_at += max_results
                else:
                    if verbose:
                        error_msg = response.text[:150] if len(response.text) > 150 else response.text
                        print(f"         ⚠️  Статус {response.status_code}: {error_msg}")
                    break
            
            if all_issues:
                if verbose:
                    print(f"         ✅ Получено через JQL: {len(all_issues)} issues")
                return all_issues
                
        except Exception as e:
            if verbose:
                print(f"         ❌ Исключение: {e}")
            continue
    
    # Метод 3: Прямой доступ к project issues (некоторые старые версии Jira)
    if verbose:
        print(f"      → Метод 3: Project Issues API")
    
    try:
        endpoint = f"{jira_url}/rest/api/2/project/{project_key}"
        response = requests.get(endpoint, headers=headers, auth=auth, timeout=30)
        
        if verbose:
            print(f"         Project API: {response.status_code}")
        
        if response.status_code == 200:
            # Проект существует, но это не дает issues напрямую
            # Пробуем простой JQL без фильтров
            endpoint = f"{jira_url}/rest/api/2/search"
            params = {
                'jql': f'key ~ "{project_key}-*"',
                'maxResults': max_results if not unlimited else 50,
                'fields': 'summary,description,comment,creator,created,updated'
            }
            
            response = requests.get(endpoint, headers=headers, auth=auth, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                issues = data.get('issues', [])
                
                if issues:
                    if verbose:
                        print(f"         ✅ Получено через key search: {len(issues)} issues")
                    return issues
                    
    except Exception as e:
        if verbose:
            print(f"         ❌ Исключение: {e}")
    
    # Если ничего не сработало
    if verbose:
        print(f"      ❌ Все методы не дали результата")
        print(f"      ℹ️  Возможные причины:")
        print(f"          - Проект пустой (нет issues)")
        print(f"          - Недостаточно прав для просмотра issues")
        print(f"          - Проект заархивирован")
        print(f"          - Специальный тип проекта")
    
    return []


def scan_issue_for_secrets(issue, patterns, jira_url):
    """
    Сканирует один issue на наличие секретов
    """
    findings = []
    issue_key = issue.get('key', 'UNKNOWN')
    issue_url = f"{jira_url}/browse/{issue_key}"
    
    fields = issue.get('fields', {})
    
    # Автор тикета
    creator = fields.get('creator', {})
    author = creator.get('displayName', 'Unknown') if creator else 'Unknown'
    author_email = creator.get('emailAddress', 'N/A') if creator else 'N/A'
    
    # Дата создания
    created = fields.get('created', 'N/A')
    
    # Summary
    summary = fields.get('summary', '')
    
    # Текст для сканирования
    texts_to_scan = []
    
    # Summary
    if summary:
        texts_to_scan.append(('Summary', summary))
    
    # Description
    description = fields.get('description')
    if description:
        # Jira может возвращать описание в разных форматах
        if isinstance(description, dict):
            # ADF (Atlassian Document Format)
            desc_text = extract_text_from_adf(description)
        else:
            desc_text = str(description)
        
        if desc_text:
            texts_to_scan.append(('Description', desc_text))
    
    # Comments
    comments = fields.get('comment', {}).get('comments', [])
    for idx, comment in enumerate(comments):
        comment_body = comment.get('body')
        if comment_body:
            if isinstance(comment_body, dict):
                comment_text = extract_text_from_adf(comment_body)
            else:
                comment_text = str(comment_body)
            
            if comment_text:
                texts_to_scan.append((f'Comment {idx+1}', comment_text))
    
    # Сканируем все тексты
    for location, text in texts_to_scan:
        secrets = scan_text_for_secrets(text, patterns)
        
        for secret in secrets:
            findings.append({
                'project_key': issue_key.split('-')[0],
                'issue_key': issue_key,
                'issue_url': issue_url,
                'summary': summary,
                'author': author,
                'author_email': author_email,
                'created': created,
                'location': location,
                'secret_type': secret['secret_type'],
                'secret_value': secret['secret_value'],
                'context': secret['context']
            })
    
    return findings


def extract_text_from_adf(adf_content):
    """
    Извлекает текст из Atlassian Document Format (ADF)
    """
    if not isinstance(adf_content, dict):
        return str(adf_content)
    
    text_parts = []
    
    def extract_recursive(node):
        if isinstance(node, dict):
            # Если есть text, добавляем его
            if 'text' in node:
                text_parts.append(node['text'])
            
            # Рекурсивно обрабатываем content
            if 'content' in node and isinstance(node['content'], list):
                for child in node['content']:
                    extract_recursive(child)
        elif isinstance(node, list):
            for item in node:
                extract_recursive(item)
    
    extract_recursive(adf_content)
    return ' '.join(text_parts)


def create_secrets_report(findings, filename="jira_secrets_report.xlsx"):
    """
    Создает Excel отчет с найденными секретами
    """
    wb = Workbook()
    
    # Лист 1: Найденные секреты
    sheet_secrets = wb.active
    sheet_secrets.title = "Found Secrets"
    
    # Заголовки
    headers = [
        'Project', 'Issue Key', 'Issue URL', 'Summary', 'Author', 'Author Email',
        'Created', 'Location', 'Secret Type', 'Secret Value', 'Context'
    ]
    
    # Стиль заголовков
    header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet_secrets.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Данные
    for row_num, finding in enumerate(findings, 2):
        sheet_secrets.cell(row=row_num, column=1).value = finding['project_key']
        sheet_secrets.cell(row=row_num, column=2).value = finding['issue_key']
        
        # URL как гиперссылка
        url_cell = sheet_secrets.cell(row=row_num, column=3)
        url_cell.value = finding['issue_url']
        url_cell.hyperlink = finding['issue_url']
        url_cell.font = Font(color='0563C1', underline='single')
        
        sheet_secrets.cell(row=row_num, column=4).value = finding['summary']
        sheet_secrets.cell(row=row_num, column=5).value = finding['author']
        sheet_secrets.cell(row=row_num, column=6).value = finding['author_email']
        sheet_secrets.cell(row=row_num, column=7).value = finding['created']
        sheet_secrets.cell(row=row_num, column=8).value = finding['location']
        sheet_secrets.cell(row=row_num, column=9).value = finding['secret_type']
        
        # Secret value - красным шрифтом
        secret_cell = sheet_secrets.cell(row=row_num, column=10)
        secret_cell.value = finding['secret_value']
        secret_cell.font = Font(color='DC143C', bold=True)
        
        sheet_secrets.cell(row=row_num, column=11).value = finding['context']
        
        # Границы для всех ячеек
        for col_num in range(1, len(headers) + 1):
            sheet_secrets.cell(row=row_num, column=col_num).border = border
    
    # Ширина колонок
    sheet_secrets.column_dimensions['A'].width = 12  # Project
    sheet_secrets.column_dimensions['B'].width = 15  # Issue Key
    sheet_secrets.column_dimensions['C'].width = 45  # URL
    sheet_secrets.column_dimensions['D'].width = 40  # Summary
    sheet_secrets.column_dimensions['E'].width = 20  # Author
    sheet_secrets.column_dimensions['F'].width = 25  # Author Email
    sheet_secrets.column_dimensions['G'].width = 20  # Created
    sheet_secrets.column_dimensions['H'].width = 15  # Location
    sheet_secrets.column_dimensions['I'].width = 30  # Secret Type
    sheet_secrets.column_dimensions['J'].width = 50  # Secret Value
    sheet_secrets.column_dimensions['K'].width = 60  # Context
    
    # Перенос текста
    for row in range(2, len(findings) + 2):
        for col in [4, 10, 11]:  # Summary, Secret, Context
            sheet_secrets.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Фиксация заголовков
    sheet_secrets.freeze_panes = 'A2'
    
    # Лист 2: Статистика
    sheet_stats = wb.create_sheet("Statistics")
    
    # Заголовок статистики
    sheet_stats['A1'] = 'Secret Scanning Statistics'
    sheet_stats['A1'].font = Font(bold=True, size=14)
    
    # Статистика
    stats_data = [
        ['Total Secrets Found:', len(findings)],
        ['Unique Secret Types:', len(set(f['secret_type'] for f in findings))],
        ['Affected Projects:', len(set(f['project_key'] for f in findings))],
        ['Affected Issues:', len(set(f['issue_key'] for f in findings))],
    ]
    
    for idx, (label, value) in enumerate(stats_data, 3):
        sheet_stats.cell(row=idx, column=1).value = label
        sheet_stats.cell(row=idx, column=1).font = Font(bold=True)
        sheet_stats.cell(row=idx, column=2).value = value
    
    # Группировка по типам секретов
    sheet_stats.cell(row=len(stats_data) + 5, column=1).value = 'Secrets by Type:'
    sheet_stats.cell(row=len(stats_data) + 5, column=1).font = Font(bold=True, size=12)
    
    secret_types = {}
    for finding in findings:
        secret_type = finding['secret_type']
        secret_types[secret_type] = secret_types.get(secret_type, 0) + 1
    
    row_offset = len(stats_data) + 6
    for idx, (secret_type, count) in enumerate(sorted(secret_types.items(), key=lambda x: x[1], reverse=True), 0):
        sheet_stats.cell(row=row_offset + idx, column=1).value = secret_type
        sheet_stats.cell(row=row_offset + idx, column=2).value = count
    
    sheet_stats.column_dimensions['A'].width = 35
    sheet_stats.column_dimensions['B'].width = 15
    
    wb.save(filename)
    return filename


def send_email_report(report_filename, findings, scan_stats, email_config):
    """
    Send email report via AWS SES
    
    Args:
        report_filename: Path to Excel report file
        findings: List of findings
        scan_stats: Dictionary with scan statistics
        email_config: Dictionary with email configuration:
            - sender: Sender email (must be verified in SES)
            - recipient: Recipient email
            - aws_region: AWS region (default: us-east-1)
    """
    if not AWS_SES_AVAILABLE:
        print("❌ AWS SES not available. Install: pip install boto3")
        return False
    
    sender = email_config.get('sender')
    recipient = email_config.get('recipient')
    aws_region = email_config.get('aws_region', 'us-east-1')
    
    if not sender or not recipient:
        print("❌ Email sender and recipient are required")
        return False
    
    try:
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Prepare subject with CRITICAL prefix
        total_secrets = len(findings)
        if total_secrets > 0:
            subject = f"CRITICAL: Jira Secrets Scanner - {total_secrets} Secret{'s' if total_secrets != 1 else ''} Found"
        else:
            subject = "INFO: Jira Secrets Scanner - No Secrets Found"
        
        # Prepare email body
        body_text = generate_email_body(findings, scan_stats)
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = recipient
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain"))
        
        # Rename report to fixed name jira_secrets.xlsx and attach
        fixed_filename = "jira_secrets.xlsx"
        if os.path.exists(report_filename):
            # Copy to fixed name
            import shutil
            shutil.copy(report_filename, fixed_filename)
            
            with open(fixed_filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {fixed_filename}",
            )
            msg.attach(part)
        
        # Send email
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=[recipient],
            RawMessage={"Data": msg.as_string()}
        )
        
        print(f"\n✅ Email sent successfully!")
        print(f"   MessageId: {response['MessageId']}")
        print(f"   From: {sender}")
        print(f"   To: {recipient}")
        print(f"   Subject: {subject}")
        return True
        
    except Exception as e:
        print(f"\n❌ Error sending email: {e}")
        return False


def generate_email_body(findings, scan_stats):
    """
    Generate email body text with scan results
    
    Args:
        findings: List of findings
        scan_stats: Dictionary with scan statistics
    
    Returns:
        String with formatted email body
    """
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    body = f"""Jira Secrets Scanner Report - {timestamp}

Summary Statistics:
* Total Secrets Found: {len(findings)}
* Affected Projects: {len(set(f['project_key'] for f in findings)) if findings else 0}
* Affected Issues: {len(set(f['issue_key'] for f in findings)) if findings else 0}

"""
    
    if findings:
        body += f"""ACTION REQUIRED:
1. Review the attached report immediately
2. Rotate/revoke exposed credentials
3. Implement proper secrets management

The detailed report is attached as XLSX file.

---
This is an automated report generated by Jira Secrets Scanner."""
    else:
        body += f"""RESULT:
No secrets detected in scanned Jira issues.
This is a good security posture!

---
This is an automated report generated by Jira Secrets Scanner."""
    
    return body


def parse_arguments():
    """
    Парсинг аргументов командной строки
    """
    parser = argparse.ArgumentParser(
        description='Сканировать Jira проекты на наличие секретов',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  # Сканировать все проекты
  python jira_secret_scanner.py --env --scan-secrets
  
  # Сканировать конкретные проекты
  python jira_secret_scanner.py -e user@company.com -t TOKEN -u URL --scan-secrets --projects PROJ1,PROJ2
  
  # Ограничить количество issues на проект
  python jira_secret_scanner.py --env --scan-secrets --max-issues 50
  
  # Указать файл с паттернами
  python jira_secret_scanner.py --env --scan-secrets --patterns custom_patterns.txt
        """
    )
    
    parser.add_argument('-e', '--email', type=str, help='Atlassian email')
    parser.add_argument('-t', '--token', type=str, help='Atlassian API token')
    parser.add_argument('-u', '--url', type=str, help='URL Jira инстанса')
    parser.add_argument('-o', '--output', type=str, help='Имя выходного файла')
    parser.add_argument('--env', action='store_true', help='Использовать .env файл')
    parser.add_argument('--env-file', type=str, default='.env', help='Путь к .env файлу')
    parser.add_argument('--scan-secrets', action='store_true', help='Сканировать issues на секреты')
    parser.add_argument('--patterns', type=str, default='secret_patterns.txt', help='Файл с паттернами секретов')
    parser.add_argument('--projects', type=str, help='Список проектов для сканирования (через запятую)')
    parser.add_argument('--max-issues', type=int, default=0, help='Максимум issues на проект (0 = все issues)')
    parser.add_argument('-q', '--quiet', action='store_true', help='Тихий режим')
    parser.add_argument('-v', '--verbose', action='store_true', help='Детальный вывод (для отладки)')
    
    # Email notification arguments (if specified, email will be sent automatically)
    parser.add_argument('--email-sender', type=str, help='Sender email address (must be verified in SES)')
    parser.add_argument('--email-recipient', type=str, help='Recipient email address')
    parser.add_argument('--aws-region', type=str, default='us-east-1', help='AWS region for SES (default: us-east-1)')
    
    return parser.parse_args()


def main():
    """
    Основная функция
    """
    args = parse_arguments()
    
    print("🔍 Jira Secret Scanner")
    print("="*80)
    
    # Загружаем .env файл
    env_vars = {}
    if args.env:
        env_vars = load_env_file(args.env_file)
        if env_vars:
            print(f"✅ Загружены данные из {args.env_file}\n")
    
    # Получаем credentials
    email = args.email or env_vars.get('JIRA_EMAIL') or input("Введите ваш Atlassian email: ").strip()
    api_token = args.token or env_vars.get('JIRA_TOKEN') or input("Введите ваш API token: ").strip()
    jira_url = args.url or env_vars.get('JIRA_URL') or input("Введите URL Jira: ").strip()
    
    if not email or not api_token or not jira_url:
        print("❌ Все поля обязательны!")
        sys.exit(1)
    
    jira_url = normalize_jira_url(jira_url)
    
    # Получаем проекты
    print("\n🔄 Получаем список проектов...")
    projects = get_jira_projects(email, api_token, jira_url)
    
    if not projects:
        print("❌ Не удалось получить список проектов")
        sys.exit(1)
    
    print(f"✅ Найдено проектов: {len(projects)}\n")
    
    # Фильтруем проекты если указаны конкретные
    if args.projects:
        selected_keys = [k.strip().upper() for k in args.projects.split(',')]
        projects = [p for p in projects if p.get('key', '').upper() in selected_keys]
        print(f"🎯 Выбрано проектов для сканирования: {len(projects)}")
    
    # Сканирование на секреты
    if args.scan_secrets:
        print("\n🔐 Начинаем сканирование на секреты...\n")
        
        # Загружаем паттерны
        patterns = load_secret_patterns(args.patterns)
        print(f"✅ Загружено паттернов: {len(patterns)}\n")
        
        all_findings = []
        total_issues_scanned = 0
        
        for project in projects:
            project_key = project.get('key', 'UNKNOWN')
            project_name = project.get('name', 'Unknown')
            
            print(f"📊 Сканирую проект: {project_key} - {project_name}")
            
            # Получаем issues
            issues = get_project_issues(email, api_token, jira_url, project_key, args.max_issues, args.verbose)
            print(f"   Получено issues: {len(issues)}")
            total_issues_scanned += len(issues)
            
            # Сканируем каждый issue
            for issue in issues:
                findings = scan_issue_for_secrets(issue, patterns, jira_url)
                if findings:
                    all_findings.extend(findings)
                    if not args.quiet:
                        print(f"   ⚠️  Найдено секретов в {issue.get('key')}: {len(findings)}")
            
            print()
        
        # Создаем отчет
        print(f"\n📊 Всего найдено секретов: {len(all_findings)}\n")
        
        # Статистика для email
        scan_stats = {
            'projects_scanned': len(projects),
            'issues_scanned': total_issues_scanned,
        }
        
        report_filename = None
        
        if all_findings:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = args.output or f"jira_secrets_report_{timestamp}.xlsx"
            
            if not report_filename.endswith('.xlsx'):
                report_filename += '.xlsx'
            
            print("📝 Создаю отчет...")
            create_secrets_report(all_findings, report_filename)
            print(f"✅ Отчет создан: {report_filename}")
            
            # Краткая статистика
            print(f"\n📈 Статистика:")
            print(f"   Затронуто проектов: {len(set(f['project_key'] for f in all_findings))}")
            print(f"   Затронуто issues: {len(set(f['issue_key'] for f in all_findings))}")
            print(f"   Типов секретов: {len(set(f['secret_type'] for f in all_findings))}")
        else:
            print("✅ Секретов не найдено!")
        
        # Отправка email если указаны sender и recipient
        if args.email_sender and args.email_recipient:
            print("\n📧 Отправка email отчета...")
            
            email_config = {
                'sender': args.email_sender,
                'recipient': args.email_recipient,
                'aws_region': args.aws_region
            }
            
            # Создаем временный отчет если не был создан (нет секретов)
            if not report_filename:
                report_filename = "jira_secrets_report_temp.xlsx"
                create_secrets_report(all_findings, report_filename)
            
            send_email_report(report_filename, all_findings, scan_stats, email_config)
    
    else:
        print("ℹ️  Используйте флаг --scan-secrets для сканирования на секреты")


if __name__ == "__main__":
    main()
